from PyQt5 import uic
from pathlib import Path
from PyQt5 import QtWidgets, QtSql, QtCore, QtGui
from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook
from openpyxl.styles import Border, Side

templates_path = Path("templates", "Отчет по выпускам.xlsx")
path = Path("venv", "ui", "ReportForm.ui")

class ReportForm(QtWidgets.QDialog):
    def __init__(self):
        super(ReportForm, self).__init__()
        uic.loadUi(path, self)
        self.setWindowModality(2)
        self.ui_init()

    def ui_init(self):
        self.dED1.dateChanged.connect(self.dateEditChanged)
        self.dED2.dateChanged.connect(self.dateEditChanged)
        self.pBSelection.clicked.connect(self.btnClickChange)
        self.tWReport.verticalHeader().setVisible(False)  # скрываем нумерацию строк в tableView
        self.tWReport.horizontalHeader().setDefaultSectionSize(100)
        self.tWReport.setColumnCount(4)
        self.tWReport.setHorizontalHeaderLabels(["ФИО пекаря", "Количество изделий", "Количество типов изделий", "Заработанная сумма"])
        self.tWReport.resizeColumnsToContents()
        self.cBSortOrder.addItem("Количество изделий")
        self.cBSortOrder.addItem("Количество типов изделий")
        self.cBSortOrder.addItem("Заработанная сумма")
        self.cBSortOrder.currentIndexChanged.connect(self.SortOrder_changed)

        self.chBDesc.stateChanged.connect(self.Desc_changed)

        self.pBClose.clicked.connect(self.btn_close_click)
        self.pBReport.clicked.connect(self.btn_report_click)

    def report_get_results(self):
        self.report_query = QtSql.QSqlQuery()
        self.report_query.prepare(
            "SELECT sotrydniki.FIO, sum(vipusk.Kolichestvo) as Kol_izdel, "
            "COUNT(izdeliya.id_izdeliya) as Type_izdel, "
            "sum(vipusk.kolichestvo * izdeliya.cena_izdeliya) as Summa "
            "FROM (sotrydniki JOIN vipusk ON (sotrydniki.id_sotrydnika = vipusk.id_sotrydnika)) " 
            "JOIN izdeliya ON (izdeliya.id_izdeliya = vipusk.id_izdeliya) "
            "WHERE (Data_vipuska >= :D1 ) and (Data_vipuska <= :D2 ) "
            "group by sotrydniki.FIO "
            f"order by {'Kol_izdel' if self.cBSortOrder.currentIndex() == 0 else 'Type_izdel' if self.cBSortOrder.currentIndex() == 1 else 'Summa'} "
            f"{'DESC' if self.chBDesc.isChecked() else 'ASC'}"
        )
        self.report_query.bindValue(":D1", self.dED1.date().toString(QtCore.Qt.ISODate))
        self.report_query.bindValue(":D2", self.dED2.date().toString(QtCore.Qt.ISODate))
        self.report_query.exec()

        self.tWReport.setRowCount(self.report_query.size())
        i = 0 #номер строки, записи
        while self.report_query.next():
            for j in range(self.tWReport.columnCount()): #j - перебираем столбцы
                item = QtWidgets.QTableWidgetItem(str(self.report_query.value(j)))
                if j==3: #анализируем только поле сумма
                    if self.report_query.value(j) > self.sBSelection.value(): #если заработанная сумма пекарем больше 360, то подкрашиваем ячейку "Сумма" в красный цвет
                        item.setBackground(QtGui.QColor(255, 0, 0, 255))
                self.tWReport.setItem(i, j, item)
            i += 1

    def dateEditChanged(self):
        self.report_get_results()

    def btnClickChange(self):
        self.report_get_results()

    def SortOrder_changed(self):
        self.report_get_results()

    def Desc_changed(self):
        self.report_get_results()

    def btn_close_click(self):
        self.close()

    def btn_report_click(self):
        #wb = load_workbook("./templates/Отчет по выпускам.xlsx") #wb - рабочая книга
        wb = load_workbook(templates_path) #открываем наш шаблон
        ws = wb["Отчет"] #указываем лист шаблона

        border = Border(left=Side(border_style="thin", color="000000"),
                        right=Side(border_style="thin", color="000000"),
                        bottom=Side(border_style="thin", color="000000"),
                        top=Side(border_style="thin", color="000000")
                        )
        ws["C3"].value = f"с {self.dED1.date().toString(QtCore.Qt.LocalDate)}" \
                         f" по {self.dED2.date().toString(QtCore.Qt.LocalDate)}"

        for i in range(self.tWReport.rowCount()):
            ws.cell(i + 6, 1).value = i+1 #для п/п в excel
            ws.cell(i + 6, 1).border = border
            for j in range(self.tWReport.columnCount()):
                if j > 0:
                    ws.cell(i+6, j+2).value = float(self.tWReport.item(i, j).text())    #в ячейку excel записываем то что в table Widget
                else:
                    ws.cell(i + 6, j + 2).value = self.tWReport.item(i, j).text()
                ws.cell(i + 6, j + 2).border = border

            # сохраняем файл с выгрузкой
        filename = QFileDialog.getSaveFileName(
            self,
            "Сохранение отчета", #заголовок окна сохранения
            "Отчет по заказам_", #имя сохраняемого документа по умолчанию
            "MS Excel Files(*.xlsx *.xls)" #в окне будут отображаться только те файлы с указанным расширением
        )

        if filename[0]: #так как необходимо сохранить только путь к файлу с именем файла, без перечисления разрешений, то выбираем первый элемент массива
            wb.save(filename[0]) #сохраняем
